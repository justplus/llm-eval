from typing import List, Dict, Any, Optional, Tuple
from app import db
from app.models import (
    ModelEvaluation, 
    ModelEvaluationDataset, 
    ModelEvaluationResult, 
    AIModel, 
    SystemDataset, 
)
from flask import current_app
import threading
from app.services.model_service import get_decrypted_api_key
from app.utils import get_beijing_time

# Evalscope imports
from evalscope.run import run_task
from evalscope.constants import JudgeStrategy
import os
import json
import pandas as pd
from datetime import datetime

# 辅助函数，尝试将evalscope的Report对象转换为可序列化的字典
def serialize_evalscope_report(report_obj):
    if hasattr(report_obj, 'to_dict') and callable(report_obj.to_dict):
        return report_obj.to_dict()
    # 如果没有to_dict，尝试直接转换，但这可能对复杂对象无效
    # 你可能需要根据ReportObject的实际结构来实现更复杂的序列化逻辑
    try:
        # 尝试使用vars()，但这只对简单对象有效
        return vars(report_obj) 
    except TypeError:
        #最后的手段，转为字符串，但这会丢失结构
        return str(report_obj) 

# 定义evalscope输出结构中review目录的名称 (基于用户提供的代码)
OUTPUTS_STRUCTURE_REVIEWS_DIR = 'reviews'

class EvaluationService:
    """模型评估服务，处理评估相关的业务逻辑"""
    
    @staticmethod
    def create_evaluation(
        user_id: int, 
        model_id: int, 
        judge_model_id: int,
        datasets: List[Dict[str, Any]],
        temperature: float, 
        max_tokens: int,
        name: Optional[str] = None,
        limit: Optional[int] = None  # 新增 limit 参数
    ) -> Optional[ModelEvaluation]:
        """
        创建一个新的模型评估任务
        """
        try:
            if not name:
                model = AIModel.query.get(model_id)
                name = f"{model.display_name if model else '未知模型'}的评估_{get_beijing_time().strftime('%Y%m%d_%H%M%S')}"
            
            evaluation = ModelEvaluation(
                user_id=user_id,
                model_id=model_id,
                judge_model_id=judge_model_id,
                name=name,
                temperature=temperature,
                max_tokens=max_tokens,
                status='pending',
                limit=limit  # 保存 limit 值
            )
            db.session.add(evaluation)
            db.session.flush()
            
            for dataset_info in datasets:
                dataset_id = dataset_info.get('dataset_id')
                eval_dataset = ModelEvaluationDataset(
                    evaluation_id=evaluation.id,
                    dataset_id=dataset_id,
                    subset=None, 
                    split=None   
                )
                db.session.add(eval_dataset)
            
            db.session.commit()
            
            app_context = current_app._get_current_object()
            threading.Thread(
                target=EvaluationService._run_evaluation_task, 
                args=(app_context, evaluation.id,)
            ).start()
            
            return evaluation
        
        except Exception as e:
            current_app.logger.error(f"创建评估任务失败: {str(e)}", exc_info=True)
            db.session.rollback()
            return None
    
    @staticmethod
    def _run_evaluation_task(app, evaluation_id: int) -> None: 
        with app.app_context(): 
            current_app.logger.info(f"[评估任务 {evaluation_id}] 开始执行。")
            evaluation = ModelEvaluation.query.get(evaluation_id)
            if not evaluation:
                current_app.logger.error(f"[评估任务 {evaluation_id}] 无法找到评估记录。")
                return
            
            evaluation.status = 'running'
            db.session.commit()
            current_app.logger.info(f"[评估任务 {evaluation_id}] 状态更新为 'running'。")
            
            model_to_evaluate = AIModel.query.get(evaluation.model_id)
            judge_model_for_evalscope = None if evaluation.judge_model_id is None else AIModel.query.get(evaluation.judge_model_id)

            if not model_to_evaluate:
                evaluation.status = 'failed'
                evaluation.result_summary = {"error": "被评估模型不存在"}
                db.session.commit()
                current_app.logger.error(f"[评估任务 {evaluation_id}] 失败: 被评估模型 ID {evaluation.model_id} 不存在。")
                return
            
            judge_api_url = None
            judge_api_key = None
            judge_model_identifier = None
            if judge_model_for_evalscope:
                judge_api_url = judge_model_for_evalscope.api_base_url
                judge_api_key = get_decrypted_api_key(judge_model_for_evalscope)
                judge_model_identifier = judge_model_for_evalscope.model_identifier

            eval_dataset_associations = ModelEvaluationDataset.query.filter_by(evaluation_id=evaluation_id).all()
            dataset_names_for_evalscope = []
            dataset_args = {}  # 新增：为自建数据集准备的dataset_args

            # 获取所有参与评估的数据集的名称 (这些是传递给evalscope的名称)
            for assoc in eval_dataset_associations:
                dataset = SystemDataset.query.get(assoc.dataset_id)
                if dataset:
                    if dataset.dataset_type == '系统':
                        # 系统数据集直接使用名称
                        dataset_names_for_evalscope.append(dataset.name) 
                    elif dataset.dataset_type == '自建':
                        # 自建数据集根据格式使用general_mcq或general_qa
                        if dataset.format == 'MCQ':
                            if 'general_mcq' not in dataset_names_for_evalscope:
                                dataset_names_for_evalscope.append('general_mcq')
                            
                            # 获取上传目录的路径
                            dataset_file_path = dataset.download_url
                            dataset_dir = os.path.dirname(dataset_file_path)
                            dataset_name = os.path.splitext(os.path.basename(dataset_file_path))[0]
                            
                            # 确保有general_mcq的dataset_args
                            if 'general_mcq' not in dataset_args:
                                dataset_args['general_mcq'] = {
                                    "local_path": dataset_dir,
                                    "subset_list": [dataset_name],
                                    'filters': {'remove_until': '</think>'} 
                                }
                            else:
                                # 如果已存在，添加到subset_list
                                if dataset_name not in dataset_args['general_mcq']['subset_list']:
                                    dataset_args['general_mcq']['subset_list'].append(dataset_name)
                            
                        elif dataset.format == 'QA':
                            if 'general_qa' not in dataset_names_for_evalscope:
                                dataset_names_for_evalscope.append('general_qa')
                            
                            # 获取上传目录的路径
                            dataset_file_path = dataset.download_url
                            dataset_dir = os.path.dirname(dataset_file_path)
                            dataset_name = os.path.splitext(os.path.basename(dataset_file_path))[0]
                            
                            # 确保有general_qa的dataset_args
                            if 'general_qa' not in dataset_args:
                                dataset_args['general_qa'] = {
                                    "local_path": dataset_dir,
                                    "subset_list": [dataset_name],
                                    'filters': {'remove_until': '</think>'} 
                                }
                            else:
                                # 如果已存在，添加到subset_list
                                if dataset_name not in dataset_args['general_qa']['subset_list']:
                                    dataset_args['general_qa']['subset_list'].append(dataset_name)
                        
                        elif dataset.format == 'FILL':
                            # 填空题格式也使用general_qa处理，因为它们都是JSONL格式的问答类型
                            if 'general_intent' not in dataset_names_for_evalscope:
                                dataset_names_for_evalscope.append('general_intent')
                            
                            # 获取上传目录的路径
                            dataset_file_path = dataset.download_url
                            dataset_dir = os.path.dirname(dataset_file_path)
                            dataset_name = os.path.splitext(os.path.basename(dataset_file_path))[0]
                            
                            # 确保有general_qa的dataset_args
                            if 'general_intent' not in dataset_args:
                                dataset_args['general_intent'] = {
                                    "local_path": dataset_dir,
                                    "subset_list": [dataset_name],
                                    'filters': {'remove_until': '</think>'}
                                }
                            else:
                                # 如果已存在，添加到subset_list
                                if dataset_name not in dataset_args['general_intent']['subset_list']:
                                    dataset_args['general_intent']['subset_list'].append(dataset_name)
                        
                        current_app.logger.info(f"[评估任务 {evaluation_id}] 添加自建数据集 {dataset.name}，格式: {dataset.format}，文件路径: {dataset.download_url}")
                else:
                    current_app.logger.warning(f"[评估任务 {evaluation_id}] 数据集ID {assoc.dataset_id} 无法找到或名称为空，已跳过。")
            
            if not dataset_names_for_evalscope:
                evaluation.status = 'failed'
                evaluation.result_summary = {"error": "没有有效的评估数据集"}
                db.session.commit()
                current_app.logger.error(f"[评估任务 {evaluation_id}] 失败: 没有提供有效的数据集进行评估。")
                return

            evalscope_run_timestamp = get_beijing_time().strftime('%Y%m%d_%H%M%S')
            base_output_dir = os.path.join('.', 'outputs', f'eval_{evaluation_id}_{evalscope_run_timestamp}') 
            try:
                os.makedirs(base_output_dir, exist_ok=True)
            except Exception as e:
                current_app.logger.error(f"[评估任务 {evaluation_id}] 创建evalscope输出目录失败: {base_output_dir}, error: {e}")
                evaluation.status = 'failed'
                evaluation.result_summary = {"error": f"创建输出目录失败: {e}"}
                db.session.commit()
                return

            decrypted_api_key = get_decrypted_api_key(model_to_evaluate)

            # 使用TaskConfig格式创建任务配置
            try:
                from evalscope import TaskConfig
            
                task_cfg_args = {
                    'eval_type': 'service', 
                    'api_url': model_to_evaluate.api_base_url,
                    'model': model_to_evaluate.model_identifier, 
                    'api_key': decrypted_api_key if decrypted_api_key else 'NO_API_KEY',
                    'datasets': dataset_names_for_evalscope, # 传递顶层数据集名称
                    'stream': True, 
                    'timeout': 12000, 
                    'work_dir': base_output_dir,
                    'eval_batch_size': 4
                }
                if judge_model_identifier:
                    task_cfg_args['judge_strategy'] = JudgeStrategy.AUTO
                    task_cfg_args['judge_model_args'] = {
                        'model_id': judge_model_identifier,
                        'api_url': judge_api_url if judge_api_url else '',
                        'api_key': judge_api_key if judge_api_key else ''
                    }
                
                # 如果有自建数据集，添加dataset_args参数
                if dataset_args:
                    task_cfg_args['dataset_args'] = dataset_args
                    current_app.logger.info(f"[评估任务 {evaluation_id}] 添加dataset_args: {json.dumps(dataset_args, indent=2)}")
                    
                if evaluation.limit and int(evaluation.limit) > 0:
                    task_cfg_args['limit'] = int(evaluation.limit)

                # 使用TaskConfig创建配置对象
                task_cfg = TaskConfig(**task_cfg_args)
                
                print(f'{judge_model_identifier} {judge_api_url} {judge_api_key}')
                current_app.logger.info(f"[评估任务 {evaluation_id}] Evalscope task_cfg: {task_cfg_args}")
            except ImportError:
                # 如果无法导入TaskConfig，则回退到使用字典
                task_cfg = {
                    'eval_type': 'service', 
                    'api_url': model_to_evaluate.api_base_url,
                    'model': model_to_evaluate.model_identifier, 
                    'api_key': decrypted_api_key if decrypted_api_key else 'NO_API_KEY',
                    'datasets': dataset_names_for_evalscope, # 传递顶层数据集名称
                    'stream': True, 
                    'timeout': 12000, 
                    'work_dir': base_output_dir,
                    'generation_config': {
                        'temperature': evaluation.temperature,
                        'max_tokens': evaluation.max_tokens,
                    },
                    'judge_model_args': { 
                        'model_id': judge_model_identifier if judge_model_identifier else '',
                        'api_url': judge_api_url if judge_api_url else '',
                        'api_key': judge_api_key if judge_api_key else ''
                    }
                }
                
                # 如果有自建数据集，添加dataset_args参数
                if dataset_args:
                    task_cfg['dataset_args'] = dataset_args
                    current_app.logger.info(f"[评估任务 {evaluation_id}] 添加dataset_args: {json.dumps(dataset_args, indent=2)}")
                    
                if evaluation.limit and int(evaluation.limit) > 0:
                    task_cfg['limit'] = int(evaluation.limit)

                print(f'{judge_model_identifier} {judge_api_url} {judge_api_key}')
                current_app.logger.info(f"[评估任务 {evaluation_id}] Evalscope task_cfg: {json.dumps(task_cfg, indent=2)}")

            evalscope_final_report = {}
            detailed_results_to_save = [] # 用于存储 ModelEvaluationResult 对象
            eval_successful = False

            try:
                raw_report_from_evalscope = run_task(task_cfg=task_cfg)
                current_app.logger.info(f"[评估任务 {evaluation_id}] Evalscope run_task completed.")
                eval_successful = True

                if isinstance(raw_report_from_evalscope, dict):
                    for ds_name_key, report_obj in raw_report_from_evalscope.items(): # ds_name_key可能包含子集信息, e.g., "cvalues/positive"
                        evalscope_final_report[ds_name_key] = serialize_evalscope_report(report_obj)
                    current_app.logger.info(f"[评估任务 {evaluation_id}] Evalscope report processed and serialized.")
                else:
                    current_app.logger.error(f"[评估任务 {evaluation_id}] Evalscope run_task did not return a dictionary as expected. Got: {type(raw_report_from_evalscope)}")
                    evalscope_final_report = {"error": "Evalscope did not return a dictionary.", "raw_output": str(raw_report_from_evalscope)}
                    eval_successful = False # 标记evalscope处理报告部分失败

                # 如果Evalscope执行成功并且我们获得了报告字典，尝试解析详细结果
                if eval_successful and isinstance(evalscope_final_report, dict) and not evalscope_final_report.get("error"):
                    current_app.logger.info(f"[评估任务 {evaluation_id}] Attempting to parse detailed review files.")
                    # 确保base_output_dir是绝对路径
                    if not os.path.isabs(base_output_dir):
                        base_output_dir = os.path.abspath(base_output_dir)
                    t_base_output_dir = base_output_dir
                    for k in os.listdir(base_output_dir):
                        t_base_output_dir = os.path.join(base_output_dir, k)
                    
                    # fix: model_to_evaluate.model_identifier可能是deepseek/deepseek-r1-0528-qwen3-8b这种格式，需要做个处理
                    t_model_identifier = model_to_evaluate.model_identifier.split('/')[-1]
                    reviews_base_path = os.path.join(t_base_output_dir, OUTPUTS_STRUCTURE_REVIEWS_DIR, t_model_identifier)
                    if os.path.isdir(reviews_base_path):
                        for review_filename_in_dir in os.listdir(reviews_base_path):
                            review_file_path = os.path.join(reviews_base_path, review_filename_in_dir)
                            
                            if os.path.isfile(review_file_path) and review_filename_in_dir.endswith('.jsonl'):
                                current_app.logger.info(f"[评估任务 {evaluation_id}] Processing review file: {review_file_path}")
                                
                                filename_stem = review_filename_in_dir[:-6] # Remove .jsonl
                                
                                try:
                                    origin_df = pd.read_json(review_file_path, lines=True)
                                    for _, item in origin_df.iterrows():
                                        raw_input = item.get('raw_input', '')
                                        raw_pred_answer = ''
                                        choices = item.get('choices', [])
                                        if choices and isinstance(choices, list) and len(choices) > 0:
                                            choice = choices[0]
                                            if isinstance(choice, dict) and 'message' in choice and isinstance(choice['message'], dict):
                                                raw_pred_answer = choice['message'].get('content', '')
                                            elif isinstance(choice, dict) and 'content' in choice:
                                                 raw_pred_answer = choice.get('content', '')
                                        
                                        review_data = {}
                                        if choices and isinstance(choices, list) and len(choices) > 0 and isinstance(choices[0], dict):
                                            review_data = choices[0].get('review', {})

                                        parsed_gold_answer = review_data.get('gold', '')
                                        parsed_pred_answer_for_feedback = review_data.get('pred', '') 
                                        score = review_data.get('result')

                                        # 处理不同格式的result
                                        try:
                                            if score is not None:
                                                if isinstance(score, dict):
                                                    # 处理复合结果格式: {"intent_result": true, "slots_result": {"miss_count": 1, "correct_count": 1, "fail_count": 0}}
                                                    if 'intent_result' in score and 'slots_result' in score:
                                                        intent_result = score.get('intent_result', False)
                                                        slots_result = score.get('slots_result', {})
                                                        
                                                        # 计算slot的F1分数
                                                        correct_count = slots_result.get('correct_count', 0)
                                                        miss_count = slots_result.get('miss_count', 0)
                                                        fail_count = slots_result.get('fail_count', 0)
                                                        
                                                        total_predicted = correct_count + fail_count
                                                        total_actual = correct_count + miss_count
                                                        
                                                        if total_predicted > 0 and total_actual > 0:
                                                            precision = correct_count / total_predicted
                                                            recall = correct_count / total_actual
                                                            if precision + recall > 0:
                                                                slot_f1 = 2 * precision * recall / (precision + recall)
                                                            else:
                                                                slot_f1 = 0.0
                                                        else:
                                                            slot_f1 = 0.0
                                                        
                                                        if correct_count + miss_count + fail_count == 0:
                                                            slot_f1 = 1.0
                                                        
                                                        # 最终分数 = intent_result * slot_f1
                                                        score = float(intent_result) * 0.5 + 0.5 * slot_f1
                                                        current_app.logger.debug(f"[评估任务 {evaluation_id}] 复合结果计算: intent={intent_result}, slot_f1={slot_f1:.4f}, final_score={score:.4f}")
                                                    else:
                                                        # 其他字典格式，尝试转换为float
                                                        score = float(score)
                                                else:
                                                    # 原有的简单数值格式
                                                    score = float(score)
                                        except (ValueError, TypeError) as e:
                                            current_app.logger.warning(f"[评估任务 {evaluation_id}] Could not parse score '{score}' for an item in {review_file_path}. Error: {str(e)}. Setting to None.")
                                            score = None

                                        result_entry = ModelEvaluationResult(
                                            evaluation_id=evaluation.id,
                                            dataset_name=filename_stem,
                                            question=str(raw_input),
                                            model_answer=str(raw_pred_answer),
                                            reference_answer=str(parsed_gold_answer),
                                            score=score,
                                            feedback=str(parsed_pred_answer_for_feedback) 
                                        )
                                        detailed_results_to_save.append(result_entry)
                                    current_app.logger.info(f"[评估任务 {evaluation_id}] Processed {len(detailed_results_to_save)} items from {review_file_path} (total for this file: {len(origin_df)})")
                                except Exception as df_exc:
                                    current_app.logger.error(f"[评估任务 {evaluation_id}] Error processing review file {review_file_path}: {str(df_exc)}", exc_info=True)
                            else:
                                current_app.logger.debug(f"[评估任务 {evaluation_id}] Skipping non-JSONL file or directory in reviews folder: {review_filename_in_dir}")
                    else:
                        current_app.logger.warning(f"[评估任务 {evaluation_id}] Reviews directory not found: {reviews_base_path}. Skipping detailed results parsing.")
                
                # 保存详细结果到数据库
                if detailed_results_to_save:
                    db.session.bulk_save_objects(detailed_results_to_save)
                    current_app.logger.info(f"[评估任务 {evaluation_id}] Saved {len(detailed_results_to_save)} detailed judge results to database.")

                evaluation.result_summary = evalscope_final_report
                evaluation.status = 'completed' if eval_successful else 'failed' # 如果evalscope执行本身就失败了，则最终状态为failed
                if eval_successful and not evalscope_final_report.get("error"): # 仅当evalscope成功且报告有效时标记完成
                     evaluation.status = 'completed'
                else: # 其他情况都算失败
                    evaluation.status = 'failed'
                    if not evalscope_final_report.get("error") and eval_successful: # evalscope执行完但报告为空或非dict
                         evalscope_final_report["error_detail"] = "Evalscope run task finished but report was invalid or empty."
                    evaluation.result_summary = evalscope_final_report


                evaluation.completed_at = get_beijing_time()
                db.session.commit() # 提交所有更改，包括状态、摘要和详细结果
                current_app.logger.info(f"[评估任务 {evaluation_id}] 评估任务处理完毕，状态: {evaluation.status}。Summary: {json.dumps(evalscope_final_report, indent=2)}")


            except Exception as es_exc:
                current_app.logger.error(f"[评估任务 {evaluation_id}] Error during evalscope execution or result processing: {str(es_exc)}", exc_info=True)
                evaluation.status = 'failed'
                evaluation.result_summary = {"error": f"Evalscope execution/processing failed: {str(es_exc)}"}
                db.session.commit() # 确保即使发生异常也提交状态
            
            finally:
                if os.path.isdir(base_output_dir):
                    try:
                        # shutil.rmtree(base_output_dir)
                        current_app.logger.info(f"[评估任务 {evaluation_id}] Successfully cleaned up evalscope output directory: {base_output_dir}")
                    except Exception as cleanup_exc:
                        current_app.logger.error(f"[评估任务 {evaluation_id}] Failed to clean up evalscope output directory {base_output_dir}: {str(cleanup_exc)}")
                else:
                    current_app.logger.warning(f"[评估任务 {evaluation_id}] Evalscope output directory not found for cleanup: {base_output_dir}")
            current_app.logger.info(f"[评估任务 {evaluation_id}] 执行线程结束。")

    @staticmethod
    def get_evaluation_by_id(evaluation_id: int, user_id: int) -> Optional[ModelEvaluation]:
        evaluation = ModelEvaluation.query.get(evaluation_id)
        if not evaluation or evaluation.user_id != user_id:
            return None
        return evaluation
    
    @staticmethod
    def get_evaluations_for_user(user_id: int, page: int = 1, per_page: int = 10) -> Tuple[List[ModelEvaluation], int]:
        query = ModelEvaluation.query.filter_by(user_id=user_id).order_by(ModelEvaluation.created_at.desc())
        total = query.count()
        evaluations = query.paginate(page=page, per_page=per_page, error_out=False).items
        return evaluations, total
    
    @staticmethod
    def get_evaluation_results(
        evaluation_id: int, 
        user_id: int, 
        page: int = 1, 
        per_page: int = 10,
        search_query: Optional[str] = None,  # 搜索查询参数
        min_score: Optional[float] = None,   # 最小分数筛选
        max_score: Optional[float] = None    # 最大分数筛选
    ) -> Tuple[List[ModelEvaluationResult], int]:
        evaluation = ModelEvaluation.query.get(evaluation_id)
        if not evaluation or evaluation.user_id != user_id:
            return [], 0
        
        query = ModelEvaluationResult.query.filter_by(evaluation_id=evaluation_id)
        
        # 如果提供了搜索查询，则添加模糊搜索条件
        if search_query:
            query = query.filter(ModelEvaluationResult.question.ilike(f"%{search_query}%"))
        
        # 添加分数范围筛选条件
        if min_score is not None:
            query = query.filter(ModelEvaluationResult.score >= min_score)
        if max_score is not None:
            query = query.filter(ModelEvaluationResult.score <= max_score)
            
        query = query.order_by(ModelEvaluationResult.id.asc())
        
        total = query.count()
        results = query.paginate(page=page, per_page=per_page, error_out=False).items
        current_app.logger.info(f"[评估结果查询] EvalID: {evaluation_id}, UserID: {user_id}, Page: {page}, Search: '{search_query}', ScoreRange: [{min_score}, {max_score}], Found: {len(results)}, Total: {total}")
        return results, total

    @staticmethod
    def export_evaluation_results_to_excel(
        evaluation_id: int, 
        user_id: int, 
        search_query: Optional[str] = None,
        min_score: Optional[float] = None,
        max_score: Optional[float] = None
    ) -> Optional[bytes]:
        """
        导出评估结果为Excel文件
        
        Args:
            evaluation_id: 评估ID
            user_id: 用户ID
            search_query: 搜索查询
            min_score: 最小分数
            max_score: 最大分数
            
        Returns:
            bytes: Excel文件的二进制数据，失败返回None
        """
        try:
            # 验证权限
            evaluation = ModelEvaluation.query.get(evaluation_id)
            if not evaluation or evaluation.user_id != user_id:
                return None
            
            # 构建查询
            query = ModelEvaluationResult.query.filter_by(evaluation_id=evaluation_id)
            
            # 应用筛选条件
            if search_query:
                query = query.filter(ModelEvaluationResult.question.ilike(f"%{search_query}%"))
            if min_score is not None:
                query = query.filter(ModelEvaluationResult.score >= min_score)
            if max_score is not None:
                query = query.filter(ModelEvaluationResult.score <= max_score)
                
            query = query.order_by(ModelEvaluationResult.id.asc())
            
            # 获取总数但不加载数据
            total_count = query.count()
            if total_count == 0:
                return None
            
            current_app.logger.info(f"开始导出评估结果，总数: {total_count}")
            
            # 创建Excel文件
            from io import BytesIO
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 分批处理数据，避免OOM
                batch_size = 100  # 每批处理1000条记录
                all_data = []
                
                # 使用yield_per进行流式查询
                for batch_start in range(0, total_count, batch_size):
                    current_app.logger.info(f"处理批次: {batch_start}-{min(batch_start + batch_size, total_count)}")
                    
                    # 分页查询当前批次
                    batch_results = query.offset(batch_start).limit(batch_size).all()
                    
                    # 处理当前批次数据
                    for idx, result in enumerate(batch_results, start=batch_start + 1):
                        # 处理问题字段，提取多轮对话
                        formatted_question = ""
                        try:
                            import json as json_lib
                            import ast
                            question_data = json_lib.loads(result.question)
                        except (json.JSONDecodeError, TypeError, ValueError):
                            question_data = ast.literal_eval(result.question)                            
                            # 构建多轮对话
                            conversation_parts = []
                            
                            # 添加历史对话
                            history = question_data.get('history') or question_data.get('hisotory', [])
                            if history:
                                for turn_idx, turn in enumerate(history, 1):
                                    if turn.get('user'):
                                        conversation_parts.append(f"用户: {turn['user']}")
                                    if turn.get('assistant'):
                                        conversation_parts.append(f"助手: {turn['assistant']}")
                            
                            # 添加当前用户问题
                            current_user_input = (question_data.get('user') or 
                                                question_data.get('question') or 
                                                question_data.get('query'))
                            if current_user_input:
                                conversation_parts.append(f"用户: {current_user_input}")
                            
                            # 组合成完整对话
                            if conversation_parts:
                                formatted_question = "\n".join(conversation_parts)
                            else:
                                formatted_question = result.question
                                
                        except (ValueError, SyntaxError, TypeError):
                            # 如果不是JSON格式或解析失败，使用原始问题
                            formatted_question = result.question
                        
                        all_data.append({
                            '序号': idx,
                            '问题': formatted_question,
                            '模型回答': result.model_answer,
                            '参考答案': result.reference_answer or '无',
                            '得分': result.score if result.score is not None else '无评分',
                            '数据集': result.dataset_name
                        })
                    
                    # 清理当前批次以释放内存
                    del batch_results
                
                # 创建DataFrame
                df = pd.DataFrame(all_data)
                current_app.logger.info(f"创建DataFrame完成，共 {len(df)} 行")
                
                # 写入Excel
                df.to_excel(writer, sheet_name='评估结果', index=False)
                
                # 获取工作表对象进行格式化
                worksheet = writer.sheets['评估结果']
                
                # 调整列宽
                column_widths = {
                    'A': 8,   # 序号
                    'B': 50,  # 问题
                    'C': 50,  # 模型回答
                    'D': 30,  # 参考答案
                    'E': 10,  # 得分
                    'F': 20   # 数据集
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 设置表头样式
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 设置数据行样式
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                # 清理DataFrame以释放内存
                del df
                del all_data
            
            output.seek(0)
            file_data = output.getvalue()
            current_app.logger.info(f"Excel导出完成，文件大小: {len(file_data)} bytes")
            return file_data
            
        except Exception as e:
            current_app.logger.error(f"导出Excel失败: {str(e)}")
            return None 